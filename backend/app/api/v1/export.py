from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import Optional
import io, csv
from datetime import datetime

from app.db.mysql_session import get_db
from app.models.mysql import Booking, User, TouristPlace, GuideProfile, BookingStatus
from app.core.dependencies import get_current_admin

router = APIRouter(prefix="/admin/export", tags=["Admin Export"])


def _get_booking_rows(db, status=None, from_date=None, to_date=None):
    q = db.query(Booking)
    if status:     q = q.filter(Booking.status == status)
    if from_date:  q = q.filter(Booking.booking_date >= from_date)
    if to_date:    q = q.filter(Booking.booking_date <= to_date)
    bookings = q.order_by(Booking.booking_date.desc()).all()

    rows = []
    for b in bookings:
        u = db.query(User).filter(User.id == b.user_id).first()
        gp = db.query(GuideProfile).filter(GuideProfile.id == b.guide_id).first()
        gu = db.query(User).filter(User.id == gp.user_id).first() if gp else None
        pl = db.query(TouristPlace).filter(TouristPlace.id == b.place_id).first()
        rows.append({
            "Booking ID": b.id,
            "Status": b.status.value,
            "Booking Date": str(b.booking_date),
            "Created At": str(b.created_at)[:19],
            "User Name": u.full_name if u else "",
            "User Email": u.email if u else "",
            "Guide Name": gu.full_name if gu else "",
            "Guide Email": gu.email if gu else "",
            "Place": pl.name if pl else "",
            "City": pl.city if pl else "",
            "State": pl.state if pl else "",
        })
    return rows


@router.get("/bookings/csv")
def export_bookings_csv(
    status: Optional[str] = Query(None),
    from_date: Optional[str] = Query(None),
    to_date: Optional[str] = Query(None),
    current_user=Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    rows = _get_booking_rows(db, status, from_date, to_date)
    if not rows:
        rows = [{}]

    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=rows[0].keys())
    writer.writeheader()
    writer.writerows(rows)

    filename = f"yatrika_bookings_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/bookings/excel")
def export_bookings_excel(
    status: Optional[str] = Query(None),
    from_date: Optional[str] = Query(None),
    to_date: Optional[str] = Query(None),
    current_user=Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        # Fall back to CSV if openpyxl not installed
        return export_bookings_csv(status, from_date, to_date, current_user, db)

    rows = _get_booking_rows(db, status, from_date, to_date)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bookings"

    header_fill = PatternFill("solid", fgColor="2E6DA4")
    header_font = Font(bold=True, color="FFFFFF")

    if rows:
        headers = list(rows[0].keys())
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[cell.column_letter].width = max(len(h) + 4, 16)

        STATUS_COLORS = {
            "completed": "C8E6C9", "accepted": "BBDEFB",
            "pending": "FFF9C4", "rejected": "FFCDD2", "cancelled": "F5F5F5",
        }
        for row_idx, row in enumerate(rows, 2):
            status_val = row.get("Status", "")
            fill_color = STATUS_COLORS.get(status_val, "FFFFFF")
            for col, val in enumerate(row.values(), 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.fill = PatternFill("solid", fgColor=fill_color)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"yatrika_bookings_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
