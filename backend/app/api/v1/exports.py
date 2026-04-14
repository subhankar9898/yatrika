"""
Admin export router — Q8: export bookings/users/guides to PDF or Excel.
Add to requirements.txt:  openpyxl==3.1.4   reportlab==4.2.2
"""
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from io import BytesIO
from datetime import datetime

from app.db.mysql_session import get_db
from app.models.mysql import Booking, User, TouristPlace, GuideProfile, BookingStatus
from app.core.dependencies import get_current_admin

router = APIRouter(prefix="/admin/export", tags=["Admin Export"])


def _get_booking_rows(db, status=None):
    q = (
        db.query(Booking, User, TouristPlace, GuideProfile)
        .join(User,          User.id          == Booking.user_id)
        .join(TouristPlace,  TouristPlace.id  == Booking.place_id)
        .join(GuideProfile,  GuideProfile.id  == Booking.guide_id)
    )
    if status:
        q = q.filter(Booking.status == status)
    return q.order_by(Booking.created_at.desc()).all()


# ─── Excel Export ─────────────────────────────────────────────────────────────

@router.get("/bookings/excel")
def export_bookings_excel(
    status: str = Query(None),
    current_user: User = Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        from fastapi import HTTPException
        raise HTTPException(500, "openpyxl not installed. Run: pip install openpyxl")

    wb = Workbook()
    ws = wb.active
    ws.title = "Bookings"

    # Header style
    header_fill = PatternFill("solid", fgColor="2E6DA4")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    headers = ["Booking ID", "User Name", "User Email", "Guide Name",
               "Place", "Date", "Status", "Created At"]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = 20

    rows = _get_booking_rows(db, status)
    for r, (booking, user, place, guide_profile) in enumerate(rows, 2):
        guide_user = db.query(User).filter(User.id == guide_profile.user_id).first()
        ws.append([
            booking.id,
            user.full_name,
            user.email,
            guide_user.full_name if guide_user else "–",
            place.name,
            str(booking.booking_date),
            booking.status.value.title(),
            str(booking.created_at)[:16],
        ])
        if r % 2 == 0:
            for col in range(1, len(headers) + 1):
                ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor="EBF5FB")

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"yatrika_bookings_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/users/excel")
def export_users_excel(
    current_user: User = Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        from fastapi import HTTPException
        raise HTTPException(500, "openpyxl not installed")

    wb = Workbook()
    ws = wb.active
    ws.title = "Users"
    header_fill = PatternFill("solid", fgColor="2E6DA4")
    header_font = Font(bold=True, color="FFFFFF")
    headers = ["ID", "Full Name", "Email", "Role", "Verified", "Active", "Joined"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        ws.column_dimensions[cell.column_letter].width = 18

    users = db.query(User).order_by(User.created_at.desc()).all()
    for u in users:
        ws.append([u.id, u.full_name, u.email, u.role.value,
                   "Yes" if u.is_verified else "No",
                   "Yes" if u.is_active else "No",
                   str(u.created_at)[:10]])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"yatrika_users_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ─── PDF Export ───────────────────────────────────────────────────────────────

@router.get("/bookings/pdf")
def export_bookings_pdf(
    status: str = Query(None),
    current_user: User = Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
    except ImportError:
        from fastapi import HTTPException
        raise HTTPException(500, "reportlab not installed. Run: pip install reportlab")

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), topMargin=30, bottomMargin=30)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph("Yatrika — Booking Report", styles["Title"]))
    elements.append(Paragraph(
        f"Generated: {datetime.now().strftime('%d %b %Y %H:%M')}  |  Filter: {status or 'All'}",
        styles["Normal"]
    ))
    elements.append(Spacer(1, 12))

    rows = _get_booking_rows(db, status)
    data = [["ID", "User", "Email", "Guide", "Place", "Date", "Status"]]
    for booking, user, place, guide_profile in rows:
        guide_user = db.query(User).filter(User.id == guide_profile.user_id).first()
        data.append([
            str(booking.id), user.full_name[:20], user.email[:25],
            (guide_user.full_name[:20] if guide_user else "–"),
            place.name[:25], str(booking.booking_date), booking.status.value.title(),
        ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2E6DA4")),
        ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",   (0, 0), (-1, 0), 9),
        ("FONTSIZE",   (0, 1), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EBF5FB")]),
        ("GRID",       (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
        ("ALIGN",      (0, 0), (-1, -1), "LEFT"),
        ("VALIGN",     (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
    ]))
    elements.append(table)
    doc.build(elements)

    buf.seek(0)
    filename = f"yatrika_bookings_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ─── Weekly Report (Last 7 Days) ─────────────────────────────────────────────

@router.get("/report/weekly")
def export_weekly_report(
    current_user: User = Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    """Download a comprehensive Excel report of the last 7 days."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        from fastapi import HTTPException
        raise HTTPException(500, "openpyxl not installed. Run: pip install openpyxl")

    from datetime import timedelta, timezone
    from app.models.mysql import UserRole, BookingStatus, GuideProfile as GP, GuidePlaceAssignment, Rating

    now = datetime.now(timezone.utc)
    seven_days_ago = now - timedelta(days=7)

    wb = Workbook()

    # ── Styling
    brand_fill = PatternFill("solid", fgColor="1E2936")
    brand_font = Font(bold=True, color="FFFFFF", size=11)
    header_border = Border(bottom=Side(style="thin", color="CCCCCC"))
    alt_fill = PatternFill("solid", fgColor="F0F4F8")
    title_font = Font(bold=True, size=14, color="1E2936")

    def style_header(ws, headers):
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.fill = brand_fill
            c.font = brand_font
            c.alignment = Alignment(horizontal="center")
            c.border = header_border
            ws.column_dimensions[c.column_letter].width = 22

    # ── Sheet 1: Summary
    ws1 = wb.active
    ws1.title = "Summary"
    ws1.cell(row=1, column=1, value="Yatrika — Weekly Report").font = title_font
    ws1.cell(row=2, column=1, value=f"Period: {seven_days_ago.strftime('%d %b %Y')} — {now.strftime('%d %b %Y')}")
    ws1.cell(row=3, column=1, value=f"Generated: {now.strftime('%d %b %Y %H:%M UTC')}")
    ws1.cell(row=5, column=1, value="Metric").font = Font(bold=True, size=11)
    ws1.cell(row=5, column=2, value="Value").font = Font(bold=True, size=11)
    ws1.column_dimensions['A'].width = 30
    ws1.column_dimensions['B'].width = 15

    total_users = db.query(User).filter(User.role == UserRole.user).count()
    total_guides = db.query(User).filter(User.role == UserRole.guide, User.is_active == True).count()
    total_places = db.query(TouristPlace).filter(TouristPlace.is_active == True).count()
    new_users_7d = db.query(User).filter(User.created_at >= seven_days_ago).count()
    bookings_7d = db.query(Booking).filter(Booking.created_at >= seven_days_ago).all()
    total_bookings_7d = len(bookings_7d)
    accepted_7d = len([b for b in bookings_7d if b.status == BookingStatus.accepted])
    completed_7d = len([b for b in bookings_7d if b.status == BookingStatus.completed])
    rejected_7d = len([b for b in bookings_7d if b.status == BookingStatus.rejected])
    pending_7d = len([b for b in bookings_7d if b.status == BookingStatus.pending])

    summary = [
        ("Total Platform Users", total_users),
        ("Total Active Guides", total_guides),
        ("Total Tourist Places", total_places),
        ("New Registrations (7d)", new_users_7d),
        ("", ""),
        ("Total Bookings (7d)", total_bookings_7d),
        ("Accepted (7d)", accepted_7d),
        ("Completed (7d)", completed_7d),
        ("Rejected (7d)", rejected_7d),
        ("Pending (7d)", pending_7d),
    ]
    for i, (metric, val) in enumerate(summary, 6):
        ws1.cell(row=i, column=1, value=metric)
        ws1.cell(row=i, column=2, value=val)
        if i % 2 == 0:
            ws1.cell(row=i, column=1).fill = alt_fill
            ws1.cell(row=i, column=2).fill = alt_fill

    # ── Sheet 2: Bookings Detail (Last 7 Days)
    ws2 = wb.create_sheet("Bookings (7d)")
    b_headers = ["ID", "User", "User Email", "Guide", "Place", "Date", "Status", "Created At"]
    style_header(ws2, b_headers)
    for r, b in enumerate(bookings_7d, 2):
        user = db.query(User).filter(User.id == b.user_id).first()
        place = db.query(TouristPlace).filter(TouristPlace.id == b.place_id).first()
        gp = db.query(GP).filter(GP.id == b.guide_id).first()
        guide_user = db.query(User).filter(User.id == gp.user_id).first() if gp else None
        ws2.append([
            b.id,
            user.full_name if user else "–",
            user.email if user else "–",
            guide_user.full_name if guide_user else "–",
            place.name if place else "–",
            str(b.booking_date),
            b.status.value.title(),
            str(b.created_at)[:16],
        ])
        if r % 2 == 0:
            for col in range(1, len(b_headers) + 1):
                ws2.cell(row=r, column=col).fill = alt_fill

    # ── Sheet 3: Daily Breakdown
    ws3 = wb.create_sheet("Daily Breakdown")
    style_header(ws3, ["Date", "Bookings Count"])
    days = {}
    for b in bookings_7d:
        day = b.created_at.strftime("%Y-%m-%d")
        days[day] = days.get(day, 0) + 1
    for r, (day, cnt) in enumerate(sorted(days.items()), 2):
        ws3.append([day, cnt])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"yatrika_weekly_report_{now.strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )

